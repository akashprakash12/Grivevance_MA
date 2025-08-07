import requests
from django.conf import settings
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone
import os
import pandas as pd
import logging
from posts.models import Post

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('facebook_realtime.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class FacebookAPI:
    def __init__(self):
        self.access_token = "EAAWySti66ogBPBoyB2Rqp3FChsOpeCNbZB6IkbFDSUMAOwuBSRIzxMdpgyVfELRWiQqM6dQ28rciJVuwBek53f3pipwUZAPZBvWsFUkOGVD5ni3MZCftGeED8ZAqfAAHbX6ZBD0keSZCT8C0ZCpenHJObyYuu4tkxG9BuwpaprAZBbXyPdpnLQpSKzDqmA9Cm41MeKe7xcnQX"
        self.page_id = settings.FB_PAGE_ID
        self.base_url = 'https://graph.facebook.com/v20.0'
        self.excel_file = os.path.join(settings.BASE_DIR, 'comments', 'comments.xlsx')
        os.makedirs(os.path.dirname(self.excel_file), exist_ok=True)
        self.max_comments_per_post = 100
        self.max_cell_chars = 32767
        self.max_retries = 3
        self.retry_delay = 5
        self.api_timeout = 30
        self.required_fields = 'id,message,created_time,permalink_url,comments.summary(true),attachments'

    def validate_token(self):
        url = f"{self.base_url}/{self.page_id}"
        params = {
            'fields': 'id,name,access_token',
            'access_token': self.access_token
        }
        response = self.make_api_request(url, params)
        if response and 'id' in response:
            logger.info("Access token validated successfully")
            return True
        logger.error(f"Access token validation failed. Response: {response}")
        return False

    def get_post_details(self, post_id):
        url = f"{self.base_url}/{post_id}"
        params = {
            'fields': self.required_fields,
            'access_token': self.access_token
        }
        return self.make_api_request(url, params)

    def make_api_request(self, url, params=None, method='GET', data=None, files=None):
        for attempt in range(self.max_retries):
            try:
                if method == 'POST':
                    response = requests.post(url, params=params, data=data, files=files, timeout=self.api_timeout)
                elif method == 'DELETE':
                    response = requests.delete(url, params=params, timeout=self.api_timeout)
                else:
                    response = requests.get(url, params=params, timeout=self.api_timeout)
                
                response.raise_for_status()
                return response.json()
            except requests.exceptions.HTTPError as e:
                error_response = {}
                try:
                    error_response = e.response.json()
                except:
                    error_response = {'error': {'message': str(e)}}
                
                logger.error(f"HTTP Error: {e}, Response: {error_response}")
                
                # Handle rate limiting
                if e.response.status_code == 429:
                    retry_after = int(e.response.headers.get('Retry-After', self.retry_delay))
                    logger.warning(f"Rate limited. Retrying after {retry_after} seconds...")
                    time.sleep(retry_after)
                    continue
                
                return error_response
            except requests.exceptions.RequestException as e:
                logger.warning(f"Attempt {attempt + 1} failed: {e}")
                time.sleep(self.retry_delay * (attempt + 1))
        
        logger.error(f"Failed to make API request after {self.max_retries} attempts")
        return {'error': {'message': 'API request failed after retries'}}

    def post_to_page(self, message=None, image_path=None):
        response = None
    
        if not image_path:
            # Text-only post
            url = f"{self.base_url}/{self.page_id}/feed"
            params = {
                'access_token': self.access_token,
                'message': message or '',
                'published': 'true'
            }
            response = self.make_api_request(url, params=params, method='POST')
        else:
            try:
                file_ext = os.path.splitext(image_path)[1].lower()
            
                if file_ext in {'.png', '.jpg', '.jpeg', '.gif'}:
                    # Photo post
                    url = f"{self.base_url}/{self.page_id}/photos"
                    params = {
                        'access_token': self.access_token,
                        'published': 'true'
                    }
                    if message:
                        params['message'] = message
                
                    with open(image_path, 'rb') as f:
                        files = {'source': f}
                        response = self.make_api_request(url, params=params, method='POST', files=files)
                
                elif file_ext in {'.mp4', '.mov', '.avi'}:
                    # Video post
                    url = f"https://graph-video.facebook.com/v20.0/{self.page_id}/videos"
                    params = {
                        'access_token': self.access_token,
                        'description': message or '',
                        'published': 'true'
                    }
                
                    with open(image_path, 'rb') as f:
                        files = {'source': f}
                        response = self.make_api_request(url, params=params, method='POST', files=files)
            except Exception as e:
                logger.error(f"Error posting media: {str(e)}", exc_info=True)
                return None
    
        if response and 'id' in response:
            # Ensure post_id is in the format PAGEID_POSTID
            if '_' not in response['id']:
                response['id'] = f"{self.page_id}_{response['id']}"
    
        return response

    def update_post(self, post_id, message, remove_media=False):
        url = f"{self.base_url}/{post_id}"
        params = {
            'access_token': self.access_token,
            'message': message
        }
        if remove_media:
            params['attached_media'] = '[]'
    
        response = self.make_api_request(url, params=params, method='POST')
    
        # Consider it successful if we get permission error (post exists but we can't access)
        if response and 'error' in response and response['error'].get('code') == 10:
            return {'success': True}
    
        return response if response else {'success': False}

    def delete_post(self, post_id):
        if '_' not in post_id:
            post_id = f"{self.page_id}_{post_id}"
        
        url = f"{self.base_url}/{post_id}"
        params = {
            'access_token': self.access_token
        }
        
        response = self.make_api_request(url, params=params, method='DELETE')
        
        # Handle special cases
        if response and 'error' in response:
            error_code = response['error'].get('code')
            # Permission error or post not found - consider these as "success" for our purposes
            if error_code in [10, 100]:
                return {'success': True}
        
        # Standard successful response
        if response and response.get('success'):
            return response
        
        # If we got here, it's a real failure
        return {
            'success': False,
            'error': response.get('error', {'message': 'Unknown error'}) if response else {'message': 'No response'}
        }

    def get_comments(self, post_id, since=None):
        if '_' not in post_id:
            post_id = f"{self.page_id}_{post_id}"

        url = f"{self.base_url}/{post_id}/comments"
        params = {
            'fields': 'id,created_time,from{name,id},message,comment_count,like_count',
            'access_token': self.access_token,
            'limit': self.max_comments_per_post,
            'order': 'chronological',
            'filter': 'stream'
        }

        if since:
            params['since'] = since.isoformat()

        comments = []
        try:
            while url:
                data = self.make_api_request(url, params if 'comments' in url else None)
                if not data or not isinstance(data, dict) or 'data' not in data:
                    break

                for comment in data.get('data', []):
                    if not isinstance(comment, dict):
                        continue
                
                    try:
                        comment_data = {
                            'id': comment.get('id', ''),
                            'user': comment.get('from', {}).get('name', 'Unknown User'),
                            'message': comment.get('message', '[No message]'),
                            'created_time': datetime.now(timezone.utc),
                            'author_id': comment.get('from', {}).get('id', ''),
                            'like_count': comment.get('like_count', 0),
                            'comment_count': comment.get('comment_count', 0),
                            'is_reply': 'parent' in comment
                        }
                    
                        if 'created_time' in comment:
                            try:
                                comment_data['created_time'] = datetime.strptime(
                                    comment['created_time'], '%Y-%m-%dT%H:%M:%S%z'
                                )
                            except ValueError as e:
                                logger.warning(f"Invalid created_time format for comment {comment.get('id')}: {str(e)}")
                    
                        comments.append(comment_data)
                    except Exception as e:
                        logger.error(f"Error processing comment {comment.get('id')}: {str(e)}")
                        continue

                url = data.get('paging', {}).get('next', None)
                time.sleep(0.5)
    
        except Exception as e:
            logger.error(f"Error fetching comments for post {post_id}: {str(e)}", exc_info=True)

        logger.info(f"Found {len(comments)} comments for post {post_id}")
        return comments

    def write_comments_to_excel(self, posts, comments_dict):
        columns = [
            'Type', 'ID', 'Author', 'Author ID', 'Time', 'Content',
            'URL', 'Parent ID', 'Parent Content', 'Likes', 'Comment Count'
        ]
        new_data = []
        
        deleted_post_ids = set(Post.objects.filter(is_deleted=True).values_list('post_id', flat=True))
        
        for post in posts:
            if post['id'] not in deleted_post_ids:
                new_data.append({
                    'Type': 'Post',
                    'ID': post['id'],
                    'Author': 'Page',
                    'Author ID': self.page_id,
                    'Time': post['created_time'],
                    'Content': post.get('message', '[No text]'),
                    'URL': post.get('permalink_url', ''),
                    'Parent ID': '',
                    'Parent Content': '',
                    'Likes': post.get('likes', {}).get('summary', {}).get('total_count', 0),
                    'Comment Count': post.get('comments', {}).get('summary', {}).get('total_count', 0)
                })
                
                for comment in comments_dict.get(post['id'], []):
                    new_data.append({
                        'Type': 'Comment',
                        'ID': comment['id'],
                        'Author': comment.get('user', 'Unknown'),
                        'Author ID': comment.get('author_id', ''),
                        'Time': comment['created_time'],
                        'Content': comment.get('message', '[No text]'),
                        'URL': f"{post.get('permalink_url', '')}?comment_id={comment['id']}",
                        'Parent ID': post['id'],
                        'Parent Content': post.get('message', '')[:50] + '...' if post.get('message') else '',
                        'Likes': comment.get('like_count', 0),
                        'Comment Count': comment.get('comment_count', 0)
                    })

        df_new = pd.DataFrame(new_data, columns=columns)
        df_new['Time'] = pd.to_datetime(df_new['Time'], errors='coerce', utc=True)
        df_new['Content'] = df_new['Content'].apply(lambda x: str(x)[:self.max_cell_chars] if pd.notna(x) else '')
        df_new['Parent Content'] = df_new['Parent Content'].apply(lambda x: str(x)[:self.max_cell_chars] if pd.notna(x) else '')

        file_exists = os.path.exists(self.excel_file)
        if file_exists:
            df_existing = pd.read_excel(self.excel_file)
            df_existing['Time'] = pd.to_datetime(df_existing['Time'], errors='coerce', utc=True)
            df_existing = df_existing[
                (~df_existing['ID'].isin(deleted_post_ids)) & 
                (~df_existing['Parent ID'].isin(deleted_post_ids))
            ]
            
            for idx, row in df_new.iterrows():
                mask = df_existing['ID'] == row['ID']
                if mask.any():
                    df_existing.loc[mask, 'Likes'] = row['Likes']
                    df_existing.loc[mask, 'Comment Count'] = row['Comment Count']
                    df_existing.loc[mask, 'Time'] = row['Time']
            
            df_combined = pd.concat([df_new, df_existing], ignore_index=True)
            df_combined = df_combined.drop_duplicates(subset=['ID'], keep='first')
            df_combined = df_combined.sort_values('Time', ascending=False, na_position='last')
            df_combined = df_combined[
                (~df_combined['ID'].isin(deleted_post_ids)) & 
                (~df_combined['Parent ID'].isin(deleted_post_ids))
            ]
        else:
            df_combined = df_new.sort_values('Time', ascending=False, na_position='last')

        df_combined['Time'] = df_combined['Time'].apply(
            lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) else ''
        )

        with pd.ExcelWriter(self.excel_file, engine='openpyxl', mode='w') as writer:
            df_combined.to_excel(writer, sheet_name='Facebook Data', index=False)
            workbook = writer.book
            worksheet = workbook['Facebook Data']
            column_widths = {
                'A': 10, 'B': 20, 'C': 15, 'D': 20, 'E': 20, 'F': 80,
                'G': 40, 'H': 20, 'I': 30, 'J': 10, 'K': 15
            }
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
                for cell in worksheet[col_letter][1:]:
                    if col_letter in ['F', 'I']:
                        cell.alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
                        cell.font = Font(name='Calibri', size=11)
                    else:
                        cell.alignment = Alignment(wrapText=False, vertical='top', horizontal='left')
                        cell.font = Font(name='Calibri', size=11)
            for cell in worksheet[1]:
                cell.alignment = Alignment(wrapText=False, vertical='center', horizontal='center')
                cell.font = Font(name='Calibri', size=11, bold=True)
            for row_idx in range(2, worksheet.max_row + 1):
                content_cell = worksheet.cell(row=row_idx, column=6)
                parent_content_cell = worksheet.cell(row=row_idx, column=9)
                max_height = 15
                for cell in [content_cell, parent_content_cell]:
                    if cell.value:
                        char_count = len(str(cell.value))
                        chars_per_line = 70 if cell.column == 6 else 40
                        lines = max(1, char_count // chars_per_line + (1 if char_count % chars_per_line else 0))
                        height = lines * 15
                        max_height = max(max_height, height)
                worksheet.row_dimensions[row_idx].height = min(max_height, 800)
        logger.info(f"Excel file updated at {self.excel_file}")