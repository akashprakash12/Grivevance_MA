import requests
from django.conf import settings
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os

class FacebookAPI:
    def __init__(self):
        self.access_token = settings.FB_ACCESS_TOKEN
        self.page_id = settings.FB_PAGE_ID
        self.base_url = 'https://graph.facebook.com/v20.0'
        self.excel_file = 'comments/comments.xlsx'

    def post_to_page(self, message=None, image_path=None):
        url = f'{self.base_url}/{self.page_id}/photos' if image_path else f'{self.base_url}/{self.page_id}/feed'
        data = {'access_token': self.access_token}
        if message:
            data['message'] = message
        files = None
        if image_path:
            files = {'source': open(image_path, 'rb')}
        response = requests.post(url, data=data, files=files)
        return response.json()

    def get_comments(self, post_id, time_limit=None):
        url = f'{self.base_url}/{post_id}/comments'
        params = {
            'access_token': self.access_token,
            'fields': 'id,from,message,created_time',
            'limit': 100
        }
        comments = []
        end_time = datetime.now() + timedelta(hours=time_limit) if time_limit else None

        while True:
            response = requests.get(url, params=params)
            data = response.json()
            if 'data' not in data:
                break
            for comment in data['data']:
                comment_time = datetime.strptime(comment['created_time'], '%Y-%m-%dT%H:%M:%S%z')
                if end_time and comment_time > end_time:
                    continue
                comments.append({
                    'id': comment['id'],
                    'user': comment['from']['name'],
                    'message': comment['message'],
                    'created_time': comment_time
                })
            if 'paging' in data and 'next' in data['paging']:
                url = data['paging']['next']
                params = {}
            else:
                break
        return comments

    def write_comments_to_excel(self, post_id, comments):
        if not os.path.exists('comments'):
            os.makedirs('comments')
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.append(['Post ID', 'Comment ID', 'User', 'Message', 'Created Time'])
            wb.save(self.excel_file)
        wb = load_workbook(self.excel_file)
        ws = wb.active
        existing_comment_ids = {row[1].value for row in ws.iter_rows(min_row=2, max_col=2)}
        for comment in comments:
            if comment['id'] not in existing_comment_ids:
                ws.append([
                    post_id,
                    comment['id'],
                    comment['user'],
                    comment['message'],
                    comment['created_time'].strftime('%Y-%m-%d %H:%M:%S')
                ])
        wb.save(self.excel_file)