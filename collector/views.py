import logging
from django import template
from django.core.exceptions import PermissionDenied
from district_officer.models import DistrictOfficerProfile  # Add this line
logger = logging.getLogger(__name__)  # Add this line
# Standard imports
import random
import re
import io
import os
import string
import smtplib
import secrets
import csv
from io import BytesIO  #
from pathlib import Path
from urllib.parse import quote
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
# Add this with your other standard imports
import pandas as pd
import openpyxl  # Required for Excel export
from fpdf import FPDF
from pathlib import Path
from django.http import JsonResponse
from .forms import AdministrativeOrderForm
from .models import CollectorOrder, CollectorOrderAssignment
from openpyxl.utils import get_column_letter  # Add this import at the top

from openpyxl import Workbook
from openpyxl.styles import Font
# Django imports
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseForbidden, FileResponse, HttpResponseRedirect
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import (
    Count, Avg, Q, F, ExpressionWrapper, DurationField, FloatField, 
    Case, When, Value
)
from django.contrib.auth.password_validation import validate_password
from django.db.models.functions import Coalesce, Cast, Round
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.timezone import is_aware
from django.utils.crypto import get_random_string
from django.conf import settings
from django.contrib.auth import get_user_model

# Date/time imports - standardized approach
from datetime import datetime, timedelta, date  # Python's datetime module
from django.utils.timezone import now as django_now  # Django's timezone-aware now

# App imports
from grievance_app.models import Grievance
from .forms import CollectorCreateUserForm, CollectorUpdateUserForm, CollectorProfileForm
from .models import CollectorProfile
from user.models import User
from core_app.models import Department, District
from officer.models import OfficerProfile
from core_app.views import auto_dept_id
from core_app.forms import DeptForm
import traceback

def _is_collector(user) -> bool:
    is_collector = CollectorProfile.objects.filter(user=user).exists()
    print("im collector for grievance excel")
    return is_collector


def _is_DO(user) -> bool:
    is_do = DistrictOfficerProfile.objects.filter(user=user).exists()
    print("im do for grievance excel")
    return is_do

### ------------------------- ###
###  Collector CRUD Functions ###
### ------------------------- ###

def create_collector(request):
    user_form = CollectorCreateUserForm(request.POST or None)
    profile_form = CollectorProfileForm(request.POST, request.FILES)

    if request.method == 'POST':
        if user_form.is_valid() and profile_form.is_valid():
            district = profile_form.cleaned_data['district']

            # Create user
            user = user_form.save(commit=False)
            user.username = auto_collector_id(district)
            user.password = make_password(user.password)
            user.user_type = 'COLLECTOR'
            user.is_active = True
            user.date_joined = timezone.now()
            user.save()

            # Create profile
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.collector_id = user.username
            profile.save()

            # Assign to 'collector' group
            collector_group = Group.objects.get(name='collector')
            user.groups.add(collector_group)

            messages.success(request, f"Collector '{user.username}' created successfully.")
            print("Uploaded file:", profile_form.cleaned_data['profile_picture'])

            return redirect('collector:view_collector')
        else:
            print("User form errors:", user_form.errors)
            print("Profile form errors:", profile_form.errors)

    return render(request, 'collector/create_collector.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })

def view_collector(request):
    collectors = CollectorProfile.objects.all()
    return render(request, 'collector/collector_list.html', {'collectors': collectors})

@login_required
def update_collector(request, username):
    """Handle collector profile updates"""

    # Ensure only collectors can access
    if request.user.user_type != "COLLECTOR":
        messages.error(request, "Only collectors may access this page")
        return HttpResponse("acess denied or something went wrong")

    user_obj = request.user
    profile_obj = get_object_or_404(CollectorProfile, user=user_obj)
    district = profile_obj.district

    if request.method == "POST":
        user_form = CollectorUpdateUserForm(request.POST, instance=user_obj)
        profile_form = CollectorProfileForm(request.POST, request.FILES, instance=profile_obj)

        # Force the district value if disabled in the form
        profile_form.data = profile_form.data.copy()
        profile_form.data["district"] = district.pk

        if user_form.is_valid() and profile_form.is_valid():
            # Save user
            user = user_form.save(commit=False)
            user.username = user.username
            user.save()

            # Save profile
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.collector_id = user.username
            profile.district = district

           
            profile.save()

            messages.success(request, "Profile updated successfully!")
            return redirect("collector:update_collector", username=user.username)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        user_form = CollectorUpdateUserForm(instance=user_obj)
        profile_form = CollectorProfileForm(instance=profile_obj)

    return render(request, "collector/update_collector.html", {
        "user_form": user_form,
        "profile_form": profile_form,
    })

def delete_collector(request, username):
    user = get_object_or_404(User, username=username, user_type='COLLECTOR')
    user.delete()
    messages.success(request, f"Collector '{username}' deleted successfully.")
    return redirect('collector:view_collector')






from django.db.models import Count, Case, When, FloatField

def rank_departments(department_queryset):
    """
    Rank departments based on grievance handling performance using a weighted point system.
    
    Args:
        department_queryset: Queryset of Department objects to rank.
    
    Returns:
        List of dictionaries containing top 3 departments with ID, name, and grievance counts.
    """
    # Aggregate grievance counts by status for each department
    dept_stats = department_queryset.annotate(
        total=Count('grievance'),
        pending_count=Count(Case(
            When(grievance__status='PENDING', then=1),
            output_field=FloatField()
        )),
        in_progress_count=Count(Case(
            When(grievance__status='IN_PROGRESS', then=1),
            output_field=FloatField()
        )),
        resolved_count=Count(Case(
            When(grievance__status='RESOLVED', then=1),
            output_field=FloatField()
        )),
        rejected_count=Count(Case(
            When(grievance__status='REJECTED', then=1),
            output_field=FloatField()
        )),
        escalated_count=Count(Case(
            When(grievance__status='ESCALATED', then=1),
            output_field=FloatField()
        ))
    ).filter(total__gt=0)

    # Weights for scoring
    weights = {
        'resolved': 1.0,      # High positive for resolved
        'in_progress': 0.3,   # Slight positive for in-progress
        'pending': -1.0,      # High negative for pending
        'rejected': -0.5,     # Moderate negative for rejected
        'escalated': -0.7     # Moderate negative for escalated
    }

    # Calculate scores and collect required fields
    dept_data = []
    for dept in dept_stats:
        # Calculate percentages for scoring
        pending_percent = (dept.pending_count / dept.total) * 100
        in_progress_percent = (dept.in_progress_count / dept.total) * 100
        resolved_percent = (dept.resolved_count / dept.total) * 100
        rejected_percent = (dept.rejected_count / dept.total) * 100
        escalated_percent = (dept.escalated_count / dept.total) * 100

        # Calculate weighted score for ranking
        score = (
            (resolved_percent * weights['resolved']) +
            (in_progress_percent * weights['in_progress']) +
            (pending_percent * weights['pending']) +
            (rejected_percent * weights['rejected']) +
            (escalated_percent * weights['escalated'])
        )

        # Collect data with counts instead of percentages
        dept_data.append({
            'code': dept.code,
            'name': dept.name,
            'total': int(dept.total),
            'pending': int(dept.pending_count),
            'in_progress': int(dept.in_progress_count),
            'resolved': int(dept.resolved_count),
            'rejected': int(dept.rejected_count),
            'escalated': int(dept.escalated_count),
            'score': score  # Used for ranking, not included in final output
        })

    # Sort by score (descending) and select top 3, excluding score
    top3 = sorted(dept_data, key=lambda x: x['score'], reverse=True)[:3]
    return [{k: v for k, v in dept.items() if k != 'score'} for dept in top3]



@login_required
def collector_dashboard(request):
    try:
        collector = (
            CollectorProfile.objects
            .select_related("district")
            .get(user=request.user)
        )
    except CollectorProfile.DoesNotExist:
        messages.error(request, "Access denied. Collector profile not found.")
        return redirect("accounts:login")

    district = collector.district

    # Get search and sort parameters
    dept_search = request.GET.get("dept_search", "").strip()
    sort_by = request.GET.get("sort", "pending")  # Default to pending

    # Base queryset
    departments_qs = (
        Department.objects
        .filter(district=district)
        .annotate(
            total=Count("grievance", filter=Q(grievance__district=district)),
            pending=Count("grievance", filter=Q(grievance__district=district, grievance__status="PENDING")),
            rejected=Count("grievance", filter=Q(grievance__district=district, grievance__status="REJECTED")),
        )
    )

    # Fuzzy match logic (only if dept_search is present)
    if dept_search:
        # First try exact match
        exact_match_qs = departments_qs.filter(
            Q(name__icontains=dept_search) | Q(code__icontains=dept_search)
        )
        if exact_match_qs.exists():
            departments_qs = exact_match_qs
        else:
            # Fuzzy search fallback
            all_departments = list(departments_qs)
            scores = []
            for d in all_departments:
                score_name = fuzz.partial_ratio(dept_search.lower(), d.name.lower())
                score_code = fuzz.partial_ratio(dept_search.lower(), d.code.lower())
                max_score = max(score_name, score_code)
                if max_score > 60:  # threshold (tune as needed)
                    scores.append((max_score, d))
            scores.sort(reverse=True)  # highest score first
            departments_qs = [d for _, d in scores]

    # Sorting logic
    if sort_by == "grievances":
        departments_qs = sorted(departments_qs, key=lambda d: d.total, reverse=True)
    else:
        departments_qs = sorted(departments_qs, key=lambda d: d.pending, reverse=True)

    # Construct final dept_data
    dept_data = []
    for d in departments_qs:
        total, pending, rejected = d.total, d.pending, d.rejected
        resolved = total - pending - rejected
        pending_pct = round((pending / total) * 100, 1) if total else 0
        rejected_pct = round((rejected / total) * 100, 1) if total else 0
        badge_class = (
            "danger" if pending_pct >= 75 else
            "warning" if pending_pct >= 25 else
            "success"
        )
        dept_data.append({
            "name": d.name,
            "code": d.code,
            "total": total,
            "pending": pending,
            "rejected": rejected,
            "resolved": resolved,
            "pending_percent": pending_pct,
            "rejected_percent": rejected_pct,
            "badge_class": badge_class,
        })

    # Top 3 departments
    top3 = rank_departments(Department.objects.filter(district=district))
    print(top3)
    # District-wide totals
    all_grievances = (
        Grievance.objects
        .filter(district=district)
        .only("grievance_id", "contact_number", "status", "subject", "due_date", "description", "applicant_address")
    )
    total_all = all_grievances.count()
    pending_all = all_grievances.filter(status="PENDING").count()
    in_progress_all = all_grievances.filter(status="IN_PROGRESS").count()
    rejected_all = all_grievances.filter(status="REJECTED").count()
    escalated_all = all_grievances.filter(status="ESCALATED").count()

    # Resolved is anything not in the above statuses (optional fallback)
    resolved_all = total_all - pending_all - in_progress_all - rejected_all - escalated_all
    
    
    district_name = district.name.lower()
    district_images_path = os.path.join(settings.STATICFILES_DIRS[0], 'images', district_name)
    district_images = []
    
    if os.path.exists(district_images_path):
        district_images = [
            f for f in os.listdir(district_images_path)
            if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))
        ]
    context = {
        "collector": collector,
        "district": district,
        "departments": dept_data,
        "top3_departments": top3,
        "counts": {
            "total": total_all,
            "pending": pending_all,
            "in_progress": in_progress_all,
            "resolved": resolved_all,
            "rejected": rejected_all,
            "escalated": escalated_all,
        },
        "all_grievances": all_grievances,
        "sort_by": sort_by,
        "dept_search": dept_search,
<<<<<<< HEAD
    }
    return render(request, "collector/collector_dashboard.html", context)

=======
        "district_images": district_images,

    }
    return render(request, "collector/collector_dashboard.html", context)



@login_required
def collector_do_profile_view(request):
    user = request.user

    if _is_collector(user):
        try:
            profile = CollectorProfile.objects.get(user=user)
            collector_profile = {
                'full_name': f"{user.first_name} {user.last_name}",
                'email': user.email,
                'username': user.username,
                'district': profile.district.name,
                'official_address': profile.official_address,
                'collector_id': profile.collector_id,
                'tenure_start': profile.tenure_start,
                'profile_picture': profile.profile_picture.url if profile.profile_picture else None,
            }
            return render(request, 'collector/profile.html', collector_profile)
        except CollectorProfile.DoesNotExist:
            messages.error(request, "Collector profile not found.")
            return redirect('collector:collector_dashboard')

    elif _is_DO(user):
        try:
            profile = DistrictOfficerProfile.objects.select_related('user', 'district').get(user=user)
            do_profile = {
                'full_name': f"{user.first_name} {user.last_name}",
                'email': user.email,
                'username': user.username,
                'district': profile.district.name,
                'officer_id': profile.officer_id,
                'assigned_on': profile.assigned_on,
                'profile_picture': profile.profile_picture.url if profile.profile_picture else None,
                "official_address":profile.official_address
            }
            return render(request, 'district_officer/profile.html', do_profile)
        except DistrictOfficerProfile.DoesNotExist:
            messages.error(request, "District Officer profile not found.")
            return redirect('district_officer:DO_dashboard')
>>>>>>> origin/main

    else:
        messages.error(request, "Access denied. Unknown role.")

@login_required
<<<<<<< HEAD
def collector_profile_view(request):
    user = request.user  # Logged-in user (from User model)

    try:
        # Get the collector profile connected to this user
        profile = CollectorProfile.objects.get(user=user)
        collector_profile = {
                    'full_name': f"{user.first_name} {user.last_name}",
                    'email': user.email,
                    'username': user.username,
                    'district': profile.district.name,
                    'official_address': profile.official_address,
                    'collector_id': profile.collector_id,
                    'tenure_start': profile.tenure_start,
                    'profile_picture': profile.profile_picture.url if profile.profile_picture else None,
                }

        return render(request, 'collector/profile.html', collector_profile)

    except CollectorProfile.DoesNotExist:
        messages.error(request, "Collector profile not found.")
        return redirect('dashboard')


@login_required
def collector_change_password(request):
=======
def change_password(request):
>>>>>>> origin/main
    """Handle password changes"""
    if request.method == "POST":
        user = request.user
        current_password = request.POST.get("old_password")
        new_password = request.POST.get("new_password")
        confirm_password = request.POST.get("confirm_password")

        if not check_password(current_password, user.password):
            messages.error(request, "Current password is incorrect")
        elif new_password != confirm_password:
            messages.error(request, "New passwords do not match")
        else:
            try:
                validate_password(new_password, user=user)
                user.set_password(new_password)
                user.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Password changed successfully!")
                if _is_collector(user):
                    return redirect("collector:update_collector", username=user.username)
                if _is_DO(user):
                    return redirect("district_officer:update_district_officer", username=user.username)
            except ValidationError as e:
                for error in e.messages:
                    messages.error(request, error)

    return render(request, "collector/password_change.html")

### ------------------------- ###
###  Department Management    ###
### ------------------------- ###

@login_required
def collector_department_create(request):
    """
    Create a new Department inside the collector's district.
    Only collectors may access this view (enforced by login_required + user type).
    """
    # ─────────────────────────────────────────────────────────────
    # 1. Fetch the collector profile for the logged‑in user
    # ─────────────────────────────────────────────────────────────
    try:
        collector = request.user.collector_profile
    except CollectorProfile.DoesNotExist:
        messages.error(request, "Collector profile not found. Please contact an administrator.")
        return redirect("collector:collector_dashboard")

    # ─────────────────────────────────────────────────────────────
    # 2. Handle POST  ➜  form submission
    # ─────────────────────────────────────────────────────────────
    if request.method == "POST":
        department_name = request.POST.get("department", "").strip()

        if not department_name:
            messages.error(request, "Department name cannot be empty.")
            return redirect("collector:collector_department_create")

        # Check uniqueness (case‑insensitive) within the same district
        if Department.objects.filter(
            name__iexact=department_name,
            district=collector.district
        ).exists():
            messages.error(
                request,
                f"A department named “{department_name}” already exists in your district."
            )
            return redirect("collector:collector_department_create")

        # Create the new department
        new_department = Department.objects.create(
            name=department_name,
            code=auto_dept_id(),
            district=collector.district,
            created_by=request.user  # ✅ Add this line
        )

        messages.success(
            request,
            f"Department “{new_department.name}” created successfully!"
        )
        return redirect("collector:collector_dashboard")

    # ─────────────────────────────────────────────────────────────
    # 3. Handle GET  ➜  show blank form
    # ─────────────────────────────────────────────────────────────
    context = {
        "auto_dept_id": auto_dept_id(),   # show next ID in the template
        "district": collector.district,
    }
    return render(request, "collector/collector_dept_create.html", context)

def department_card_view(request, department_id):
    # ---------- basic context ----------
    if _is_collector(request.user):
        collector = get_object_or_404(
            CollectorProfile.objects.select_related('district'),
            user=request.user
        )
        district = collector.district
        template='collector/each_dept.html'
    if _is_DO(request.user):
        do=get_object_or_404(DistrictOfficerProfile.objects.select_related('district'),user=request.user)
        district=do.district
        print("helooooooo")
        template='district_officer/each_dept.html'

    department  = get_object_or_404(Department, code=department_id)
    
    # ---------- filters from request ----------
    status_filter = request.GET.get("status", "ALL")
    date_from     = request.GET.get("date_from")
    date_to       = request.GET.get("date_to")
    search_query  = request.GET.get("search", "").strip()

    # ---------- start with dept + district ----------
    base_qs = Grievance.objects.filter(
        district=district,
        department=department
    )

    # ----- date range -----
    if date_from:
        base_qs = base_qs.filter(date_filed_date_gte=date_from)
    if date_to:
        base_qs = base_qs.filter(date_filed_date_lte=date_to)

    # ----- free‑text search (ID or phone) -----
    if search_query:
        base_qs = base_qs.filter(
            Q(grievance_id__icontains=search_query) |
            Q(contact_number__icontains=search_query)
        )

    # ---------- counts BEFORE status filter ----------
    counts = {
        "total":     base_qs.count(),
        "pending":   base_qs.filter(status="PENDING").count(),
        "in_progress": base_qs.filter(status="IN_PROGRESS").count(),
        "resolved":  base_qs.filter(status="RESOLVED").count(),
        "rejected":  base_qs.filter(status="REJECTED").count(),
        "escalated": base_qs.filter(status="ESCALATED").count(),
    }

    # ---------- now apply status filter for table ----------
    if status_filter != "ALL":
        grievances_qs = base_qs.filter(status=status_filter)
    else:
        grievances_qs = base_qs

    grievances_qs = grievances_qs.order_by("-date_filed")

    # ---------- pagination ----------
    paginator   = Paginator(grievances_qs, 25)
    page_number = request.GET.get("page")
    page_obj    = paginator.get_page(page_number)

    # ---------- render ----------
    context = {
        "district":       district,
        "department":     department,
        "grievances":     page_obj,
        "counts":         counts,
        "status_filter":  status_filter,
        "date_from":      date_from,
        "date_to":        date_to,
        "search_query":   search_query,
    }
    return render(request, template, context)



def update_remark(request):
    if request.method == 'POST':
        grievance_id = request.POST.get('grievance_id')
        remark = request.POST.get('remark')
        priority = request.POST.get('priority')
        next_url = request.POST.get('next')  # Get the next URL from the form

        # Fetch the grievance once
        grievance = get_object_or_404(Grievance, id=grievance_id)

        # Validate inputs
        if not remark:
            messages.error(request, "Remark is required.")
            return redirect(next_url or 'collector:collector_dashboard')

        # Update remark
        grievance.remark = remark

        # Update priority only if provided and valid
        valid_priorities = ['LOW', 'MEDIUM', 'HIGH', 'CRITICAL']
        if priority and priority in valid_priorities:
            grievance.priority = priority
        # Else, retain the existing priority (no change)

        grievance.save()

        messages.success(request, f"Remark and priority updated for GRV ID: {grievance.grievance_id}")

        # Redirect to the next URL if provided, otherwise fallback to department_card
        if next_url:
            return redirect(next_url)
        return redirect('collector:department_card', grievance.department.code)
    else:
        messages.error(request, "Invalid request method.")
        return redirect('collector:collector_dashboard')
    
    
### ------------------------- ###
###  Reporting & Export Views ###
### ------------------------- ###
@login_required
def grievance_report_view(request):
    print(f"[DEBUG] grievance_report_view accessed by {request.user.username}")
    
    user_type = request.user.user_type.lower()  # Normalize casing
    print(f"[DEBUG] user_type: {user_type}")

    district = None

    try:
        if user_type == "collector":
            profile = get_object_or_404(CollectorProfile, user=request.user)
            district = profile.district
            print(f"[DEBUG] Collector district: {district}")
        elif user_type == "district_officer":
            profile = get_object_or_404(DistrictOfficerProfile, user=request.user)
            district = profile.district
            print(f"[DEBUG] District Officer district: {district}")
        else:
            print(f"[DEBUG] Invalid user type: {user_type}")
            raise PermissionDenied("You don't have permission to view this report")

        departments = Department.objects.filter(district=district)
        print(f"[DEBUG] Departments count: {departments.count()}")

        grievances = _filtered_grievance_qs(request, district)
        print(f"[DEBUG] Grievances count: {grievances.count()}")

        if _is_collector(request.user):
            template = "collector/grievance_report.html"
        elif _is_DO(request.user):
            template = "district_officer/grievance_report.html"
        else:
            print("[DEBUG] User matched no known role (collector/DO)")
            return HttpResponseForbidden("Access Denied: You do not have permission to view this page.")

        print(f"[DEBUG] Using template: {template}")

        return render(request, template, {
            "grievances": grievances,
            "departments": departments,
            "district": district
        })

    except Exception as e:
        print(f"[ERROR] grievance_report_view failed: {e}")
        return HttpResponseForbidden("Something went wrong or access denied.")
@login_required
def department_report_view(request):
    user_type = request.user.user_type
    district = None
    print("hello",request.user)
    print("user_type is:", user_type)

    if user_type == "COLLECTOR":
        user = get_object_or_404(CollectorProfile, user=request.user)
        district = user.district
        template="collector/department_report.html"
        print("hello im coll")

    elif user_type == "district_officer":  # Changed this condition
        user = get_object_or_404(DistrictOfficerProfile, user=request.user)
        district = user.district
        template="district_officer/department_report.html"
        print("hello im do")

    else:
        # Handle other user types or raise permission denied
        raise PermissionDenied("You don't have permission to view this report")


    qs = Department.objects.filter(district=district)
    search = request.GET.get("search")
    sort_by = request.GET.get("sort_by")

    if search:
        qs = qs.filter(name__icontains=search)

    report_data = []
    for dept in qs:
        grievances = Grievance.objects.filter(department=dept, district=district)
        total = grievances.count()
        pending = grievances.filter(status="PENDING").count()
        resolved = grievances.filter(status="RESOLVED").count()
        rejected = grievances.filter(status="REJECTED").count()
        escalated = grievances.filter(status="ESCALATED").count()

        resolution_rate = round((resolved / total) * 100, 2) if total else 0

        report_data.append({
            "name": dept.name,
            "code": dept.code,
            "total": total,
            "pending": pending,
            "resolved": resolved,
            "rejected": rejected,
            "escalated": escalated,
            "resolution_rate": resolution_rate,
        })

    if sort_by:
        reverse = sort_by.startswith("-")
        key = sort_by.lstrip("-")
        report_data.sort(key=lambda x: x.get(key, 0), reverse=reverse)

    # ✅ Store final filtered and sorted data for export
    request.session['filtered_department_report'] = report_data

    # paginator = Paginator(report_data, 10)
    # page_obj = paginator.get_page(request.GET.get("page"))

    total_depts = qs.count()
    avg_resolution_rate = round(sum(d["resolution_rate"] for d in report_data) / total_depts, 2) if total_depts else 0

    context = {
        "district": district,
        "department_report": report_data,
        "department_count": total_depts,
        "summary": {
            "total_depts": total_depts,
            "avg_resolution_rate": avg_resolution_rate,
        }
    }
    return render(request, template, context)

@login_required
def export_grievance_excel(request):
    # Get the collector based on logged-in user
    if _is_collector(request.user):
        collector = get_object_or_404(CollectorProfile, user=request.user)
        district = collector.district
    elif _is_DO(request.user):
        do = get_object_or_404(DistrictOfficerProfile, user=request.user)
        district = do.district
    else:
        HttpResponse("permission denied")
    # Get filtered grievance queryset and select relevant fields
    qs = _filtered_grievance_qs(request, district).values(
        'grievance_id',
        'date_filed',
        'last_updated',
        'subject',
        'description',
        'source',
        'status',
        'priority',
        'due_date',
        'applicant_name',
        'applicant_address',
        'contact_number',
        'email',
        'department__name',
        'district__name'
    )

    # Clean and format date fields
    for row in qs:
        for date_field in ['date_filed', 'last_updated', 'due_date']:
            field_value = row.get(date_field)
            if field_value:
                if isinstance(field_value, datetime):
                    if is_aware(field_value):
                        field_value = field_value.replace(tzinfo=None)
                    row[date_field] = field_value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(field_value, date):
                    row[date_field] = field_value.strftime('%Y-%m-%d')

    # Create a DataFrame
    df = pd.DataFrame(qs)

    # Rename columns for Excel headers
    df.columns = [
        'Grievance ID',
        'Date Filed',
        'Last Updated',
        'Subject',
        'Description',
        'Source',
        'Status',
        'Priority',
        'Due Date',
        'Applicant Name',
        'Applicant Address',
        'Contact Number',
        'Email',
        'Department',
        'District'
    ]

    # Write to Excel in memory
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Grievances", index=False)

        # Auto-format Excel worksheet
        worksheet = writer.sheets['Grievances']
        column_widths = {
            'A': 15, 'B': 20, 'C': 20, 'D': 30, 'E': 50,
            'F': 15, 'G': 15, 'H': 10, 'I': 20, 'J': 25,
            'K': 40, 'L': 15, 'M': 25, 'N': 25, 'O': 20
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions

    buffer.seek(0)

    # Return Excel file as response
    return FileResponse(
        buffer,
        as_attachment=True,
        filename=f"{district.code.lower()}_grievance_report.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@login_required
def export_grievance_pdf(request):
    try:
        if _is_collector(request.user):
            collector = get_object_or_404(CollectorProfile, user=request.user)
            district = collector.district
        elif _is_DO(request.user):
            do = get_object_or_404(DistrictOfficerProfile, user=request.user)
            district = do.district
        else:
            HttpResponse("permission denied")
        grievances = _filtered_grievance_qs(request, district)

        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.set_margins(left=10, top=15, right=10)
        pdf.add_page()

        # Title
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, f"{district.name.upper()} - Detailed Grievance Report", ln=1, align='C')
        pdf.ln(3)

        # Column headers and widths
        headers = ["ID", "Date", "Status", "Applicant", "Contact", "Address", "Subject"]
        widths = [28, 22, 25, 35, 28, 55, 60]  # Total ~253mm for landscape A4 (297mm wide minus margins)

        def print_table_header():
            pdf.set_font('Arial', 'B', 9)
            for header, width in zip(headers, widths):
                pdf.cell(width, 7, header, border=1, align='C')
            pdf.ln()

        print_table_header()
        pdf.set_font('Arial', '', 8)
        row_height = 5

        for grievance in grievances:
            data = [
                grievance.grievance_id,
                grievance.date_filed.strftime('%d-%m-%Y') if grievance.date_filed else "N/A",
                grievance.status or "N/A",
                grievance.applicant_name or "N/A",
                grievance.contact_number or "N/A",
                grievance.applicant_address or "N/A",
                grievance.subject or "N/A",
            ]

            # Estimate number of lines per cell
            cell_lines = []
            max_lines = 1
            for i, (text, width) in enumerate(zip(data, widths)):
                lines = pdf.multi_cell(width, row_height, str(text), border=0, split_only=True)
                cell_lines.append(lines)
                max_lines = max(max_lines, len(lines))
            total_height = row_height * max_lines

            # Page break check
            if pdf.get_y() + total_height > 190:
                pdf.add_page()
                print_table_header()
                pdf.set_font('Arial', '', 8)

            # Draw cells
            y_start = pdf.get_y()
            x_start = pdf.get_x()

            for i, width in enumerate(widths):
                pdf.set_xy(x_start, y_start)
                if i in [5, 6]:  # Multiline cells (Address, Subject)
                    pdf.multi_cell(width, row_height, data[i], border=1)
                else:
                    pdf.cell(width, total_height, data[i], border=1, ln=0)
                x_start += width

            pdf.set_y(y_start + total_height)

        # Output
        buffer = io.BytesIO()
        pdf.output(buffer)
        buffer.seek(0)

        return FileResponse(buffer, as_attachment=True,
                            filename=f"{district.code}_detailed_grievances.pdf",
                            content_type="application/pdf")

    except Exception as e:
        logger.error(f"PDF export error: {str(e)}", exc_info=True)
        messages.error(request, "Failed to export report. Please try again.")
        return redirect("collector:grievance_report")
@login_required
def export_department_excel(request):
    
    if _is_collector(request.user):
        collector = get_object_or_404(CollectorProfile, user=request.user)
        district = collector.district
    elif _is_DO(request.user):
        do = get_object_or_404(DistrictOfficerProfile, user=request.user)
        district = do.district

    else:
        HttpResponse("permission denied")

    report_data = request.session.get("filtered_department_report")

    # Fallback: if session is missing
    if not report_data:
        qs = Department.objects.filter(district=district)
        report_data = []
        for dept in qs:
            grievances = Grievance.objects.filter(department=dept, district=district)
            total = grievances.count()
            resolved = grievances.filter(status="RESOLVED").count()
            report_data.append({
                "name": dept.name,
                "code": dept.code,
                "total": total,
                "pending": grievances.filter(status="PENDING").count(),
                "resolved": resolved,
                "rejected": grievances.filter(status="REJECTED").count(),
                "escalated": grievances.filter(status="ESCALATED").count(),
                "resolution_rate": round((resolved / total) * 100, 2) if total else 0,
            })

    df = (
        pd.DataFrame(report_data)
          .rename(columns={
              "name": "Department",
              "code": "Code",
              "total": "Total Grievances",
              "pending": "Pending",
              "resolved": "Resolved",
              "rejected": "Rejected",
              "escalated": "Escalated",
              "resolution_rate": "Resolution Rate (%)",
          })
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Department Report", index=False)
        ws = writer.sheets["Department Report"]

        widths = [25, 15, 18, 12, 12, 12, 12, 20, 25]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    buffer.seek(0)
    return FileResponse(
        buffer,
        as_attachment=True,
        filename=f"{district.name.replace(' ', '_')}_department_report.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@login_required
def export_department_pdf(request):
    try:
        if _is_collector(request.user):
            collector = get_object_or_404(CollectorProfile, user=request.user)
            district = collector.district
        elif _is_DO(request.user):
            do = get_object_or_404(DistrictOfficerProfile, user=request.user)
            district = do.district

        else:
            HttpResponse("permission denied")

        report_data = request.session.get("filtered_department_report", [])

        if not report_data:
            messages.error(request, "No filtered data available to export. Please generate a report first.")
            return redirect('department_report_view')

        # --- PDF Setup ---
        pdf = FPDF('L', 'mm', 'A4')  # Landscape A4
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        try:
            font_dir = Path(settings.BASE_DIR) / "fonts"
            pdf.add_font('DejaVu', '', str(font_dir / 'DejaVuSans.ttf'), uni=True)
            pdf.add_font('DejaVu', 'B', str(font_dir / 'DejaVuSans-Bold.ttf'), uni=True)
            font = 'DejaVu'
        except:
            font = 'Arial'

        pdf.set_font(font, 'B', 16)
        pdf.cell(0, 10, f'{district.name} - Department Performance Report', 0, 1, 'C')
        pdf.ln(5)

        # --- Table Header ---
        pdf.set_fill_color(79, 129, 189)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font(font, 'B', 10)

        col_widths = [60, 22, 22, 22, 22, 22, 30, 35]
        headers = [
            'Department', 'Code', 'Total', 'Pending',
            'Resolved', 'Rejected', 'Resolution %',
        ]

        for header, width in zip(headers, col_widths):
            pdf.cell(width, 8, header, border=1, align='C', fill=True)
        pdf.ln()

        # --- Table Body ---
        pdf.set_text_color(0, 0, 0)
        pdf.set_font(font, '', 9)

        for dept in report_data:
            pdf.cell(col_widths[0], 8, dept['name'], border=1)
            pdf.cell(col_widths[1], 8, dept['code'], border=1, align='C')
            pdf.cell(col_widths[2], 8, str(dept['total']), border=1, align='C')
            pdf.cell(col_widths[3], 8, str(dept['pending']), border=1, align='C')
            pdf.cell(col_widths[4], 8, str(dept['resolved']), border=1, align='C')
            pdf.cell(col_widths[5], 8, str(dept['rejected']), border=1, align='C')
            pdf.cell(col_widths[6], 8, f"{dept['resolution_rate']}%", border=1, align='C')
            pdf.ln()

        # --- Output the PDF ---
        buffer = io.BytesIO()
        pdf.output(buffer)
        buffer.seek(0)

        return FileResponse(
            buffer,
            as_attachment=True,
            filename=f"{district.name.replace(' ', '_')}_department_report.pdf",
            content_type='application/pdf'
        )

    except Exception as e:
        messages.error(request, f"Failed to generate PDF: {str(e)}")
        return redirect('department_report_view')

@login_required
def details_download(request, grievance_id):
    try:
        if _is_collector(request.user):
            collector = get_object_or_404(CollectorProfile, user=request.user)
            district = collector.district
        elif _is_DO(request.user):
            do = get_object_or_404(DistrictOfficerProfile, user=request.user)
            district = do.district
        else:
            HttpResponse("permission denied")

        grievance = get_object_or_404(
            Grievance.objects.select_related("department", "district"),
            grievance_id=grievance_id,
            district=district
        )

        # PDF Setup
        pdf = FPDF("P", "mm", "A4")
        margin = 15
        pdf.set_auto_page_break(auto=True, margin=margin)
        pdf.set_margins(left=margin, top=margin, right=margin)
        pdf.add_page()
        pdf.set_line_width(0.5)
        pdf.rect(margin - 5, margin - 5, 210 - (margin - 5)*2, 297 - (margin - 5)*2)

        try:
            font_dir = Path(settings.BASE_DIR) / "fonts"
            pdf.add_font("DV", "", str(font_dir / "DejaVuSans.ttf"), uni=True)
            pdf.add_font("DV", "B", str(font_dir / "DejaVuSans-Bold.ttf"), uni=True)
            font_family = "DV"
        except:
            font_family = "Arial"
        # Title
        label_width = 35
        value_width = 150

        # Grievance ID
        pdf.set_font(font_family, "B", 13)
        pdf.cell(0, 8, f"Grievance ID: {grievance.grievance_id}", ln=True)

        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Applicant:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.applicant_name, ln=True)

        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Address:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.applicant_address, ln=True)

        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Contact:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.contact_number, ln=True)

        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Email:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.email, ln=True)

        pdf.ln(2)

        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Subject:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.subject, ln=True)

        if grievance.description:
            pdf.set_font(font_family, "B", 12)
            pdf.cell(0, 6, "Description:", ln=True)
            pdf.set_font(font_family, "", 12)
            pdf.multi_cell(0, 6, grievance.description)

        
        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.source, ln=True)
        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Status:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.status, ln=True)
        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Source:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.source, ln=True)

        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Priority:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.priority, ln=True)

        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Department:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, grievance.department.name, ln=True)

        pdf.cell(0, 6, "-" * 50, ln=True)

        # Dates
        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Filed:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, f"{grievance.date_filed:%Y-%m-%d}", ln=True)

        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Last Updated:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, f"{grievance.last_updated:%Y-%m-%d}", ln=True)

        pdf.set_font(font_family, "B", 12)
        pdf.cell(label_width, 6, "Due:", ln=False)
        pdf.set_font(font_family, "", 12)
        pdf.cell(value_width, 6, f"{grievance.due_date:%Y-%m-%d}", ln=True)



        # Return PDF
        buffer = io.BytesIO()
        pdf.output(buffer)
        buffer.seek(0)

        return FileResponse(
            buffer,
            as_attachment=True,
            filename=f"{grievance.grievance_id}_report.pdf",
            content_type="application/pdf"
        )

    except Exception as e:
        messages.error(request, f"Failed to generate PDF: {str(e)}")
        return redirect("collector:grievance_report")


@login_required
def department_grievances_download(request, department_id):
    try:
        # 1. Security/ownership checks
        if _is_collector(request.user):
            collector = get_object_or_404(CollectorProfile, user=request.user)
            district = collector.district
        if _is_DO(request.user):
            do=get_object_or_404(DistrictOfficerProfile,user=request.user)
            district = do.district

        department = get_object_or_404(Department, code=department_id, district=district)

        # 2. Apply filters from request
        qs = (
            Grievance.objects
            .filter(district=district, department=department)
            .select_related("department")
            .order_by("-date_filed")
        )

        status = request.GET.get("status", "ALL")
        if status != "ALL":
            qs = qs.filter(status=status)

        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from:
            qs = qs.filter(date_filed__gte=date_from)
        if date_to:
            qs = qs.filter(date_filed__lte=date_to)

        search = request.GET.get("search")
        if search:
            qs = qs.filter(
                Q(grievance_id__icontains=search) |
                Q(contact_number__icontains=search)
            )

        # 3. PDF setup with simplified font handling
        pdf = FPDF("P", "mm", "A4")
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Use built-in fonts to avoid font registration issues
        font_family = 'Arial'  # Default built-in font
        
        # Alternatively, if you really need DejaVu, register all variants properly:
        try:
            font_dir = Path(settings.BASE_DIR) / "fonts"
            if font_dir.exists():
                # Register all required variants
                pdf.add_font('DejaVu', '', str(font_dir / 'DejaVuSans.ttf'), uni=True)
                pdf.add_font('DejaVu', 'B', str(font_dir / 'DejaVuSans-Bold.ttf'), uni=True)
                pdf.add_font('DejaVu', 'I', str(font_dir / 'DejaVuSans-Oblique.ttf'), uni=True)
                pdf.add_font('DejaVu', 'BI', str(font_dir / 'DejaVuSans-BoldOblique.ttf'), uni=True)
                font_family = 'DejaVu'
        except Exception as font_error:
            print(f"Font loading error, falling back to Arial: {font_error}")
            font_family = 'Arial'

        # Header
        pdf.set_font(font_family, "B", 16)
        pdf.cell(0, 10, f"{district.name} District", ln=True, align="C")
        pdf.cell(0, 10, f"{department.name} - Grievance Report", ln=True, align="C")
        
        # Filters info
        pdf.set_font(font_family, "", 10)  # Note: Using regular style, not italic
        filter_text = []
        if status != "ALL":
            filter_text.append(f"Status: {status}")
        if date_from:
            filter_text.append(f"From: {date_from}")
        if date_to:
            filter_text.append(f"To: {date_to}")
        if search:
            filter_text.append(f"Search: {search}")
        
        if filter_text:
            pdf.ln(5)
            pdf.cell(0, 6, "Filters: " + " | ".join(filter_text), ln=True)
        
        pdf.ln(10)

        # Grievance list
        if qs.exists():
            pdf.set_font(font_family, "B", 11)
            for grievance in qs:
                pdf.cell(0, 8, f"GRV {grievance.grievance_id}  -  {grievance.applicant_name}", ln=True)
                
                pdf.set_font(font_family, "", 9)  # Regular style for details
                filed_date = grievance.date_filed.strftime('%Y-%m-%d %H:%M')
                due_date = grievance.due_date.strftime('%Y-%m-%d') if grievance.due_date else 'Not set'
                
                pdf.multi_cell(
                    0,
                    5,
                    (
                        f"Contact: {grievance.contact_number} | Priority: {grievance.priority} | Status: {grievance.get_status_display()}\n"
                        f"Subject: {grievance.subject or '-'}\n"
                        f"Description: {grievance.description}\n"
                        f"Filed: {filed_date} | Due: {due_date}"
                    ),
                )
                
                pdf.set_draw_color(200, 200, 200)
                pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                pdf.ln(5)
                pdf.set_font(font_family, "B", 11)
        else:
            pdf.set_font(font_family, "", 12)
            pdf.cell(0, 10, "No grievances found matching the selected filters", ln=True, align="C")

        # Footer - using regular font instead of italic to avoid issues
        pdf.set_y(-15)
        pdf.set_font(font_family, "", 8)  # Changed from "I" to ""
        pdf.cell(0, 10, f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M')}", 0, 0, 'C')

        # 4. Stream PDF to browser
        buffer = io.BytesIO()
        pdf.output(buffer)
        buffer.seek(0)

        file_name = f"{department.code.lower()}grievances{timezone.now().strftime('%Y%m%d')}.pdf"
        return FileResponse(buffer, as_attachment=True,
                          filename=file_name,
                          content_type="application/pdf")

    except Exception as e:
        print(f"Error in PDF generation: {str(e)}")
        return HttpResponse(f"An error occurred while generating the PDF: {str(e)}", status=500)

@login_required
def export_department_grievances_excel(request, department_id):
    """
    Export department grievances to Excel with filters preserved
    """
    try:
        if _is_collector(request.user):
            collector = get_object_or_404(CollectorProfile, user=request.user)
            district = collector.district
            department = get_object_or_404(Department, code=department_id, district=district)

        if _is_DO(request.user):
            do=get_object_or_404(DistrictOfficerProfile,user=request.user)
            district = do.district
            department = get_object_or_404(Department, code=department_id, district=district)

        
        # 2. Apply filters (same as department_card_view)
        grievances = Grievance.objects.filter(
            district=district,
            department=department
        ).select_related('department')
        
        status_filter = request.GET.get("status", "ALL")
        if status_filter != "ALL":
            grievances = grievances.filter(status=status_filter)
            
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from:
            grievances = grievances.filter(date_filed__gte=date_from)
        if date_to:
            grievances = grievances.filter(date_filed__lte=date_to)
            
        search_query = request.GET.get("search", "").strip()
        if search_query:
            grievances = grievances.filter(
                Q(grievance_id__icontains=search_query) |
                Q(contact_number__icontains=search_query)
            )
            
        # 3. Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Grievances"
        
        # Headers
        headers = [
            "Grievance ID", "Date Filed", "Status", "Priority",
            "Applicant Name", "Contact", "Address", "Subject", 
            "Description", "Due Date", "Last Updated"
        ]
        ws.append(headers)
        
        # Data rows
        for grievance in grievances:
            ws.append([
                grievance.grievance_id,
                grievance.date_filed.strftime('%Y-%m-%d %H:%M') if grievance.date_filed else "",
                grievance.get_status_display(),
                grievance.get_priority_display(),
                grievance.applicant_name,
                grievance.contact_number,
                grievance.applicant_address,
                grievance.subject,
                grievance.description,
                grievance.due_date.strftime('%Y-%m-%d') if grievance.due_date else "",
                grievance.last_updated.strftime('%Y-%m-%d %H:%M') if grievance.last_updated else "",
            ])
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Column widths
        column_widths = [18, 16, 12, 12, 25, 15, 30, 30, 50, 12, 16]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
        # 4. Return the file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"{department.code}_grievances_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        messages.error(request, f"Failed to generate Excel: {str(e)}")
        return redirect('collector:department_card', department_id=department_id)
### ------------------------- ###
###  Communication Views      ###
### ------------------------- ###



def collector_or_do_check(user):
    return _is_collector(user) or _is_DO(user)

@login_required
def officer_details(request):
    # Get user profile and district
    if _is_collector(request.user):
        profile = get_object_or_404(CollectorProfile, user=request.user)
        template = "collector/hod_list.html"
    else:
        profile = get_object_or_404(DistrictOfficerProfile, user=request.user)
        template = "district_officer/hod_list.html"
    
    district = profile.district

    # Get HOD officers in the district
    hod_officers = OfficerProfile.objects.select_related('user', 'department') \
        .filter(is_hod=True, department__district=district)
    
    # Get active DO for the district (if any)
    do_details = DistrictOfficerProfile.objects.filter(
        district=district, 
        is_active=True
    ).first()
    collector_details=CollectorProfile.objects.filter(
        district=district, 
        is_active=True
    ).first()
    context = {
<<<<<<< HEAD
=======
        "profile": profile,
        "do_details": do_details,
>>>>>>> origin/main
        'hod_officers': hod_officers,
        "collector_details": collector_details,  # Only relevant in DO portal

        'district': district,  # Pass district object
        'is_collector': _is_collector(request.user),
    }
    return render(request, template, context)

@login_required
def send_email_redirect(request, officer_email):
    if not officer_email or '@' not in officer_email:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
    
    # Gmail Compose URL
    compose_url = f"https://mail.google.com/mail/?view=cm&fs=1&tf=1&to={quote(officer_email)}"
    
    # Add subject if needed
    if hasattr(request.user, 'collectorprofile'):
        subject = f"Regarding {request.user.collectorprofile.district.name} District Matters"
        compose_url += f"&su={quote(subject)}"
    
    return HttpResponseRedirect(compose_url)











### ------------------------- ###
###  Helper Functions         ###
### ------------------------- ###

def auto_collector_id(district):
    prefix = 'COLL'
    district_code = district.code.upper()

    # Check both CollectorProfile.collector_id AND User.username
    # to ensure no duplicates in either table
    existing_collector_ids = CollectorProfile.objects.filter(
        district=district
    ).values_list('collector_id', flat=True)

    existing_usernames = User.objects.filter(
        username__startswith=prefix,
        username__endswith=district_code
    ).values_list('username', flat=True)

    max_number = 0
    pattern = re.compile(rf"{prefix}(\d+){district_code}$", re.IGNORECASE)

    # Check collector profiles
    for cid in existing_collector_ids:
        match = pattern.match(cid)
        if match:
            num = int(match.group(1))
            if num > max_number:
                max_number = num

    # Check usernames (important!)
    for username in existing_usernames:
        match = pattern.match(username)
        if match:
            num = int(match.group(1))
            if num > max_number:
                max_number = num

    next_number = max_number + 1
    return f"{prefix}{next_number}{district_code}"

def _filtered_grievance_qs(request, district):
    """
    Return Grievances in the given district filtered by query params
    """
    # Get cleaned parameters
    status = (request.GET.get("status") or "ALL").strip().upper()
    dept_code = (request.GET.get("department") or "").strip()
    search = (request.GET.get("search") or "").strip()
    date_from = parse_date((request.GET.get("date_from") or "").strip())
    date_to = parse_date((request.GET.get("date_to") or "").strip())

    # Base queryset
    qs = (
        Grievance.objects
        .filter(district=district)
        .select_related("department", "district")
        .order_by("-date_filed")
    )

    # Status filter
    if status != "ALL":
        qs = qs.filter(status__iexact=status)

    # Department filter - using code field
    if dept_code:
        qs = qs.filter(department__code__iexact=dept_code)  # Fixed this line


    # Search filter
    if search:
        search_q = Q()
        for term in search.split():
            search_q |= (
                Q(grievance_id__icontains=term) |
                Q(contact_number__icontains=term) |
                Q(email__icontains=term) |
                Q(applicant_name__icontains=term)
            )
        qs = qs.filter(search_q)

    # Date range filter
   # Date range filter
    date_filter = Q()
    if date_from:
        date_filter &= Q(date_filed__date__gte=date_from)  # Also fixed date lookup
    if date_to:
        date_filter &= Q(date_filed__date__lte=date_to)  # Also fixed date lookup
    if date_filter:
        qs = qs.filter(date_filter)


    return qs

def validate_email(email):
    """Basic email validation"""
    return re.match(r'^[^@]+@[^@]+\.[^@]+$', email) is not None


# Constants
OTP_EXPIRY_MINUTES = 5
MAX_ATTEMPTS = 3
COOLDOWN_MINUTES = 1

@login_required
def collector_handover_otp(request):
    """
    Handles OTP generation and verification for collector handover
    """
    if request.method == "POST":
        action = request.POST.get("action")

        # Generate OTP
        if action == "generate":
            # Check cooldown status
            cooldown_until = request.session.get("otp_cooldown_until")
            if cooldown_until and django_now() < datetime.fromisoformat(cooldown_until):
                remaining = (datetime.fromisoformat(cooldown_until) - django_now()).seconds
                messages.error(request, f"Please wait {remaining} seconds before retrying.")
                return render(request, "collector/handover.html", {"step": "generate"})

            # Generate and store OTP
            otp = str(random.randint(100_000, 999_999))
            request.session.update({
                "handover_otp": otp,
                "otp_generated_at": django_now().isoformat(),
                "otp_attempts": 0
            })

            # Send OTP email
            try:
                send_mail(
        subject=f"Collector Role Handover Verification – {request.user.collector_profile.district.name}",
        message=f"""\
    Dear {request.user.first_name},

    You are initiating the official handover of your Collector responsibilities for the district of {request.user.collector_profile.district.name}.

    To proceed with this process, please use the secure one-time verification code provided below:

    One-Time Password (OTP): {otp}  
    Valid Until: {(django_now() + timedelta(minutes=OTP_EXPIRY_MINUTES)).strftime('%I:%M %p on %d %b %Y')}

    ------------------------------------------------------------
    Important Security Notice:
    - This OTP grants full access to the role transfer process.
    - Do not share this code with anyone under any circumstances.
    - The Grievance Redressal System team will never request this code via email, phone, or any other communication channel.
    - If you did not initiate this request, please report the activity immediately.

    ------------------------------------------------------------
    Post-Verification Process:
    - A new Collector account will be generated upon successful verification.
    - Your current administrative access will be permanently revoked.
    - Confirmation emails will be sent to both involved parties following the completion of the process.

    Disclaimer:  
    Unauthorized use of this code, including intentional or accidental sharing, will be considered a serious breach of protocol. The current account holder will be held fully responsible for any consequences, as per the Digital Governance Security Policy.

    ------------------------------------------------------------
    For safety-related queries or assistance, please contact GULBEE:

    Security Desk:+91 9999988888  
    Email: security@gulbee-system.org


    Thank you for your dedicated service.

    Sincerely,  
    Office of Digital Governance  
    Grievance Redressal System  
    https://portal.grievance-system.org
    """,

                    from_email="Grievance System Security <no-reply@grievance-system.org>",
                    recipient_list=[request.user.email],
                )
                messages.success(request, "Verification code sent to your email.")
            except Exception as e:
                messages.error(request, "Failed to send verification code. Please try again.")
                return render(request, "collector/handover.html", {"step": "generate"})

            return render(request, "collector/handover.html", {
                "step": "verify",
                "otp_expiry_minutes": OTP_EXPIRY_MINUTES
            })

        # Verify OTP
        elif action == "verify":
            # Validate OTP existence
            session_otp = request.session.get("handover_otp")
            generated_at = request.session.get("otp_generated_at")
            if not session_otp or not generated_at:
                messages.error(request, "Verification code expired. Please request a new one.")
                return redirect("collector:collector_handover_otp")

            # Check OTP expiration
            expiry_time = datetime.fromisoformat(generated_at) + timedelta(minutes=OTP_EXPIRY_MINUTES)
            if django_now() > expiry_time:
                request.session.pop("handover_otp", None)
                request.session.pop("otp_generated_at", None)
                messages.error(request, "Verification code expired. Please request a new one.")
                return redirect("collector:collector_handover_otp")

            # Check attempt limit
            attempts = request.session.get("otp_attempts", 0) + 1
            if attempts > MAX_ATTEMPTS:
                cooldown_until = (django_now() + timedelta(minutes=COOLDOWN_MINUTES)).isoformat()
                request.session["otp_cooldown_until"] = cooldown_until
                request.session.pop("handover_otp", None)
                request.session.pop("otp_generated_at", None)
                messages.error(request, f"Too many attempts. Please wait {COOLDOWN_MINUTES} minute(s).")
                return redirect("collector:collector_handover_otp")

            # Verify OTP
            if request.POST.get("otp") == session_otp:
                request.session.pop("handover_otp", None)
                request.session.pop("otp_generated_at", None)
                request.session["otp_verified"] = True
                return redirect("collector:new_collector_info")
            else:
                request.session["otp_attempts"] = attempts
                messages.error(request, f"Invalid code. {MAX_ATTEMPTS-attempts} attempts remaining.")
                return render(request, "collector/handover.html", {
                    "step": "verify",
                    "otp_expiry_minutes": OTP_EXPIRY_MINUTES
                })

    return render(request, "collector/handover.html", {"step": "generate"})

@login_required
def new_collector_info(request):
    """
    Collect new collector's basic info after OTP verification
    and complete the handover process with new collector OTP verification
    """
    if not request.session.get('otp_verified'):
        messages.error(request, "Verification required first")
        return redirect('collector:collector_handover_otp')
    
    current_collector = request.user.collector_profile
    district = current_collector.district
    
    if request.method == 'POST':
        if request.session.get('new_collector_data') and 'new_collector_otp' in request.POST:
            return _verify_new_collector_otp(request)
        return _process_new_collector_info(request, current_collector, district)
    
    if request.session.get('new_collector_data'):
        return render(request, 'collector/verify_new_collector_otp.html')
    
    return render(request, 'collector/new_collector_info.html')

def _process_new_collector_info(request, current_collector, district):
    """Handle the initial form submission"""
    first_name = request.POST.get('first_name', '').strip()
    last_name = request.POST.get('last_name', '').strip()
    email = request.POST.get('email', '').strip().lower()
    phone = request.POST.get('phone', '').strip()
    
    if not all([first_name, last_name, email]):
        messages.error(request, "First name, last name and email are required")
        return render(request, 'collector/new_collector_info.html', {
            'first_name': first_name,
            'last_name': last_name,
            'email': email,
            'phone': phone,
        })
    
    otp = str(random.randint(100000, 999999))
    request.session['new_collector_data'] = {
        'first_name': first_name,
        'last_name': last_name,
        'email': email,
        'phone': phone,
        'otp': otp,
        'otp_created_at': timezone.now().isoformat(),
    }

    try:
        send_mail(
    subject="Confirm Your Appointment as Collector – Grievance Redressal System",
    message=f"""\
Dear {first_name} {last_name},

We are pleased to inform you that you have been selected for the esteemed position of Collector in the district of {district.name}, under the Grievance Redressal System.

To complete your appointment, please verify your identity using the one-time security code below:

Verification Code: {otp}  
Code Expiry: 10 minutes from the time this email was sent

------------------------------------------------------------
Next Steps:
- Your official Collector account will be created upon successful verification.
- Login credentials will be sent to your registered email address.
- You will be required to set a new password upon your first login.

------------------------------------------------------------
Security Advisory:
- This code is strictly confidential.
- Do not share this code with anyone under any circumstances.
- If you did not expect this communication or are unaware of this nomination, please disregard this message and report any suspicious activity to the support team.

------------------------------------------------------------
We look forward to your leadership and contribution to the governance framework.
For safety-related queries or assistance, please contact GULBEE:

    Security Desk:+91 9999988888  
    Email: security@gulbee-system.org
Sincerely,  
Grievance Redressal Team  
noreply@grievance-system.com
""",


            from_email="noreply@grievance-system.com",
            recipient_list=[email],
        )
        messages.info(request, "Verification code sent to the new collector.")
        return redirect('collector:new_collector_info')
    except Exception as e:
        messages.error(request, "Failed to send verification code.")
        return render(request, 'collector/new_collector_info.html', {
            'first_name': first_name,
            'last_name': last_name,
            'email': email,
            'phone': phone,
        })
def _verify_new_collector_otp(request):
    """Verify the OTP from new collector and complete handover"""
    session_data = request.session.get('new_collector_data')
    if not session_data:
        messages.error(request, "Session expired. Please start over.")
        return redirect('collector:collector_handover_otp')
    
    # Check OTP expiry (10 minutes)
    otp_created_at = datetime.fromisoformat(session_data['otp_created_at'])
    if django_now() - otp_created_at > timedelta(minutes=10):
        messages.error(request, "Verification code expired.")
        request.session.pop('new_collector_data', None)
        return redirect('collector:collector_handover_otp')
    
    # Verify OTP
    submitted_otp = request.POST.get('new_collector_otp', '').strip()
    if submitted_otp != session_data['otp']:
        messages.error(request, "Invalid verification code.")
        return render(request, 'collector/verify_new_collector_otp.html')

    try:
        with transaction.atomic():
            # Generate credentials
            password = get_random_string(12)
            district = request.user.collector_profile.district
            
            # Create new user
            new_user = User.objects.create(
                first_name=session_data['first_name'],
                last_name=session_data['last_name'],
                email=session_data['email'],
                phone=session_data['phone'] if session_data['phone'] else None,
                username=auto_collector_id(district),
                password=make_password(password),
                user_type='COLLECTOR',
                is_active=True,
                date_joined=django_now()
            )
            
            # Assign to collector group
            collector_group, _ = Group.objects.get_or_create(name='collector')
            new_user.groups.add(collector_group)
            
            # Create collector profile
            CollectorProfile.objects.create(
                user=new_user,
                district=district,
                collector_id=new_user.username,
                tenure_start=django_now().date(),
                official_address="NA",
                profile_picture=None,
                is_active=True
            )
            
            # Deactivate current collector
            current_user = request.user
            current_user.is_active = False
            current_user.save()
            
            # Deactivate current collector profile
            try:
                current_collector = CollectorProfile.objects.get(user=current_user)
                current_collector.is_active = False
                current_collector.save()
            except CollectorProfile.DoesNotExist:
                pass
            
            # Clear user session
            from django.contrib.auth import logout
            logout(request)
            
# Send notification emails
            try:
                # Email to new collector
                send_mail(
    subject="Welcome to the Grievance Redressal System – Collector Credentials Enclosed",
    message=f"""\
Dear {new_user.first_name},

Congratulations on your appointment as the Collector of {district.name}.

This position carries a significant responsibility in fostering transparency, ensuring accountability, and delivering timely resolution to public grievances within your jurisdiction.

------------------------------------------------------------
Your Login Credentials:
- Username: {new_user.username}
- Temporary Password: {password}
- Login Portal: https://grievance-system.example.com/login

------------------------------------------------------------
Next Steps:
1. Log in and change your password immediately.
2. Update your profile information and preferences.
3. Begin overseeing grievance management and district-level operations.

------------------------------------------------------------
Learn More:
Gain an understanding of our system’s structure, role hierarchy, and operational protocols:  
https://grievance-system.example.com/about

------------------------------------------------------------
Security Notice:
- This communication is confidential and intended solely for the recipient.
- All system activities and transitions are monitored and logged.
- Sharing this email or its contents is strictly prohibited and will result in full accountability for any resulting misuse.
- Unauthorized use or disclosure may lead to disciplinary action and legal consequences.

------------------------------------------------------------
We value your leadership and look forward to your service in advancing public governance.

Wishing you success in this important role.
For safety-related queries or assistance, please contact GULBEE:

    Security Desk:+91 9999988888  
    Email: security@gulbee-system.org
Sincerely,  
Grievance Redressal System Team  
support@example.com | +1 (555) 123-4567
"""
,

                    from_email="noreply@grievance-system.com",
                    recipient_list=[new_user.email],
                )

                # Email to current collector
                send_mail(
    subject="Collector Handover Completed – Acknowledgment of Your Service",
    message=f"""\
Dear {current_user.first_name},

This is to confirm that the handover of the Collector responsibilities for the district of {district.name} has been successfully completed.  
Your duties have now been formally transitioned to Mr./Ms. {new_user.first_name} {new_user.last_name}.

------------------------------------------------------------
Acknowledgment of Your Service:
We sincerely thank you for your valuable contributions throughout your tenure.  
Your leadership and dedication have made a significant and lasting impact on the grievance redressal framework of the district.

------------------------------------------------------------
Important Notice:
As part of this secure transition, your access to the system has been permanently revoked.

Disclaimer:  
Any misuse, mishandling, or unauthorized activity involving your credentials during the transition period will be considered a breach of the system's security policy.  
In such cases, you may be held fully accountable and subject to disciplinary or legal action as applicable.

------------------------------------------------------------
Review Your Service Record:
To view a summary of your actions and contributions during your tenure, please refer to the report below:  
https://grievance-system.example.com/contribution-report?user={current_user.username}

------------------------------------------------------------
We extend our best wishes for your continued success in your future endeavors.  
Your service has been greatly valued and appreciated.

For safety-related queries or assistance, please contact GULBEE:

    Security Desk:+91 9999988888  
    Email: security@gulbee-system.org
Sincerely,  
Grievance Redressal System Team  
noreply@grievance-system.com
"""
,

                    from_email="noreply@grievance-system.com",
                    recipient_list=[current_user.email],
    )

            except Exception as email_error:
                logger.error(f"Failed to send handover emails: {str(email_error)}")
            
            # Clear all session data
            request.session.flush()
            
            messages.success(request, "Handover completed successfully! The new collector can now login.")
            return redirect('accounts:login')
            
    except Exception as e:
        logger.error(f"Handover failed: {str(e)}")
        messages.error(request, "Error during handover process. Please contact support.")
        return redirect('collector:collector_handover_otp')
def handover_complete(request):
    messages.success(request, "Handover completed successfully!")
    return redirect('accounts:login')


User = get_user_model()


def _get_district_officer_email(district):
    """
    Adjust to your real model: we assume one active DistrictOfficerProfile per district.
    """
    from district_officer.models import DistrictOfficerProfile  # or wherever it lives
    officer = DistrictOfficerProfile.objects.filter(district=district, is_active=True).first()
    return officer.user.email if officer else None


# ─────────────────────────────────────────────────────────────
@login_required
def collector_forgot_password(request):
    """
    Step 1 – Logged-in collector clicks 'Forgot Password', triggers OTPs to self + DO.
    """
    user = request.user

    # Safety check
    if user.user_type != "COLLECTOR":
        messages.error(request, "Access denied.")
        return redirect("accounts:dashboard")  # or wherever

    district = user.collector_profile.district
    officer_email = _get_district_officer_email(district)
    if not officer_email:
        messages.error(request, "District officer email not configured.")
        return render(request, "collector/forgot_password.html")

    # Generate OTPs
    collector_otp = f"{random.randint(100000, 999999)}"
    officer_otp   = f"{random.randint(100000, 999999)}"

    # Save to session
    request.session["fp_data"] = {
        "collector_user_id": user.id,
        "collector_otp": collector_otp,
        "officer_otp": officer_otp,
    }

    # Send OTPs
    # Send OTP to collector
    send_mail(
    subject="Your Password Reset OTP – Confidential",
    message=f"""\
Dear {user.first_name},

You have requested to reset your password for the Collector Portal.  
Please use the following One-Time Password (OTP) to proceed:

OTP: {collector_otp}

------------------------------------------------------------
Security Advisory:
- This OTP is strictly confidential.
- Do not share this code with anyone under any circumstances.
- Any unauthorized usage or sharing of this code will be fully tracked.
- You will be held solely responsible for any activity performed using your credentials.

This OTP is valid for a limited period. Kindly enter it promptly to complete your password reset.

If you did not initiate this request, please contact support immediately to secure your account.

------------------------------------------------------------
All password reset activities are monitored and logged to ensure system integrity and your personal security.

For safety-related queries or assistance, please contact GULBEE:

    Security Desk:+91 9999988888  
    Email: security@gulbee-system.org
Sincerely,  
Grievance Redressal System Team  
noreply@grievance-system.com
"""
,

        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
    )

    # Send OTP to District Officer
    send_mail(
    subject="Collector Password Reset – OTP Approval Required",
    message=f"""\
Dear District Officer,

A password reset request has been initiated by the Collector of {district.name}.  
In accordance with security protocols, your approval is required to authorize this action.

Please find below your One-Time Password (OTP) for approval:

Approval OTP: {officer_otp}

------------------------------------------------------------
Confidentiality Notice:
- This OTP must be communicated **only** to the concerned Collector.
- Do **not** share or forward this code to anyone else under any circumstances.
- All actions during this password reset process are monitored and securely logged.
- Any misuse, unauthorized sharing, or breach will result in full accountability and may lead to disciplinary or legal consequences.

By providing this OTP, you affirm that the request has been verified and approved by you.

If you believe this request is suspicious or unauthorized, please report it immediately to the system administrator.

------------------------------------------------------------
Thank you for upholding the security and integrity of district-level operations.

For safety-related queries or assistance, please contact GULBEE:

    Security Desk:+91 9999988888  
    Email: security@gulbee-system.org
Sincerely,  
Grievance Redressal System Team  
noreply@grievance-system.com
"""
,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[officer_email],
    )


    messages.info(request, "OTPs sent to you and your district officer.")
    return redirect("collector:collector_verify_two_otps")

# ─────────────────────────────────────────────────────────────
def collector_verify_two_otps(request):
    """
    Step 2 – user enters both OTPs
    """
    data = request.session.get("fp_data")
    if not data:
        messages.error(request, "Session expired — start again.")
        return redirect("collector:collector_forgot_password")

    if request.method == "POST":
        coll_otp = request.POST.get("collector_otp", "").strip()
        off_otp  = request.POST.get("officer_otp", "").strip()

        if coll_otp == data["collector_otp"] and off_otp == data["officer_otp"]:
            request.session["fp_verified"] = True
            return redirect("collector:collector_reset_password")
        messages.error(request, "OTP mismatch. Try again.")

    return render(request, "collector/verify_two_otps.html")


# ─────────────────────────────────────────────────────────────
def collector_reset_password(request):
    """
    Step 3 – set new password
    """
    if not request.session.get("fp_verified"):
        return redirect("collector:collector_forgot_password")

    data = request.session.get("fp_data")
    try:
        user = User.objects.get(id=data["collector_user_id"])
    except User.DoesNotExist:
        messages.error(request, "Collector not found.")
        return redirect("collector:collector_forgot_password")

    if request.method == "POST":
        p1 = request.POST.get("password1")
        p2 = request.POST.get("password2")

        if not p1 or not p2:
            messages.error(request, "Please fill both fields.")
        elif p1 != p2:
            messages.error(request, "Passwords do not match.")
        elif len(p1) < 8:
            messages.error(request, "Password must be at least 8 characters.")
        else:
            user.password = make_password(p1)
            user.save()

            # Clean up session
            for key in ("fp_data", "fp_verified"):
                request.session.pop(key, None)

            messages.success(request, "Password reset successful. Please log in.")
            return redirect("accounts:login")

    return render(request, "collector/reset_password.html")    
<<<<<<< HEAD
=======




# @login_required
# def create_collector_order(request):
#     try:
#         collector = get_object_or_404(CollectorProfile, user=request.user)
#         district = collector.district
#     except CollectorProfile.DoesNotExist:
#         return redirect("collector:collector_dashboard")

#     departments = Department.objects.filter(district=district).order_by('name')

#     if request.method == "POST":
#         form = AdministrativeOrderForm(request.POST, request.FILES, user=request.user)
#         if form.is_valid():
#             order = form.save(commit=False)
#             order.assigned_by = request.user
#             order.created_at = timezone.now()

#             if not order.assigned_officer:
#                 selected_depts = form.cleaned_data['departments']
#                 hods = OfficerProfile.objects.filter(department__in=selected_depts, is_hod=True)
#                 if hods.exists():
#                     order.assigned_officer = hods.first()

#             order.save()
#             form.save_m2m()
#             return redirect("collector:collector_dashboard")
#     else:
#         form = AdministrativeOrderForm(user=request.user)

#     return render(request, "collector/create_order.html", {"form": form, "departments": departments})


# @login_required
# def get_officers_by_department(request):
#     dept_ids = request.GET.getlist('dept_ids[]')
#     officers = OfficerProfile.objects.filter(
#         department_id__in=dept_ids, is_hod=True
#     ).select_related('user', 'department')

#     officer_list = [
#         {
#             "id": officer.id,
#             "name": f"{officer.user.first_name} {officer.user.last_name} ({officer.department.name})"
#         }
#         for officer in officers
#     ]
#     return JsonResponse({"officers": officer_list})







#for collector ordering


# @login_required
# def create_collector_order(request):
#     try:
#         collector = get_object_or_404(CollectorProfile, user=request.user)
#         district = collector.district
#     except CollectorProfile.DoesNotExist:
#         return redirect("collector:collector_dashboard")

#     departments = Department.objects.filter(district=district).order_by("name")

#     if request.method == "POST":
#         title = request.POST.get("title")
#         remark = request.POST.get("remark")
#         due_date = request.POST.get("due_date")
#         attachment = request.FILES.get("attachment")

#         # Validate required fields
#         if not (title and remark and due_date and request.POST.getlist("departments[]")):
#             return render(
#                 request,
#                 "collector/create_order.html",
#                 {
#                     "departments": departments,
#                     "error": "All required fields must be filled, and at least one department must be selected."
#                 }
#             )

#         # Create the order
#         order = CollectorOrder.objects.create(
#             title=title,
#             remark=remark,
#             due_date=due_date,
#             attachment=attachment,
#             assigned_by=request.user,
#             created_at=timezone.now()
#         )

#         # Process departments and officers
#         dept_ids = request.POST.getlist("departments[]")
#         officer_ids = request.POST.getlist("officers[]")

#         for dept_id in dept_ids:
#             try:
#                 department = Department.objects.get(code=dept_id, district=district)
#                 order.departments.add(department)

#                 # If no officers selected, assign to HOD
#                 if not officer_ids:
#                     try:
#                         hod = OfficerProfile.objects.get(department=department, is_hod=True)
#                         CollectorOrderAssignment.objects.create(order=order, officer=hod)
#                     except OfficerProfile.DoesNotExist:
#                         # Skip if no HOD exists for the department
#                         continue
#                 else:
#                     # Assign to selected officers
#                     for officer_id in officer_ids:
#                         try:
#                             officer = OfficerProfile.objects.get(id=officer_id, department__district=district)
#                             CollectorOrderAssignment.objects.create(order=order, officer=officer)
#                         except OfficerProfile.DoesNotExist:
#                             continue  # Skip invalid officer IDs
#             except Department.DoesNotExist:
#                 continue  # Skip invalid department IDs

#         return redirect("collector:collector_dashboard")

#     return render(request, "collector/create_order.html", {"departments": departments})

# @login_required
# def get_officers_by_department(request):
#     dept_ids = request.GET.getlist('dept_ids[]')
#     try:
#         collector = get_object_or_404(CollectorProfile, user=request.user)
#         officers = OfficerProfile.objects.filter(
#             department__code__in=dept_ids,
#             department__district=collector.district
#         ).select_related('user', 'department')

#         officer_list = [
#             {
#                 "id": officer.id,
#                 "name": f"{officer.user.first_name} {officer.user.last_name} ({officer.department.name})"
#             }
#             for officer in officers
#         ]
#         return JsonResponse({"officers": officer_list})
#     except CollectorProfile.DoesNotExist:
#         return JsonResponse({"officers": []}, status=403)
>>>>>>> origin/main
